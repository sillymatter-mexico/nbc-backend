from universal.celery_app import app
from users.models import ReportUsers, ClientUser, Session
from users.controllers import ClientUserControllers
from openpyxl import Workbook
from django.db.models import Q
from django.conf import settings
from django.http.response import HttpResponse
from django.core.mail import send_mail, EmailMessage
from django.db.models import Avg, Count, Sum
from users.models import UpdateFile
from users.controllers import ClientUserControllers
from datetime import timedelta

@app.task
def ReportClientUser(data):
    report = ReportUsers.objects.get(pk=data['id'])
    start = report.start_date
    finish = report.finish_date
    start = str(start)+' 00:00:01+00:00'
    finish = str(finish) + ' 23:59:59+00:00'
    cont = 0
    query_set= Q(enable=True, created__range=[start, finish])
    list_user = ClientUser.objects.filter(query_set)
    total = len(list_user)
    email = report.email
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'ID Uuiversal'
    ws['B1'] = 'Fecha de resgistro'
    ws['C1'] = 'Total puntuación juegos'
    ws['D1'] = 'Juegos completados'
    ws['E1'] = 'Acumulaciones premier'
    ws['F1'] = 'Puntuacion Total'
    cont2 = 2
    for user in list_user:
        cont = cont + 1
        ws.cell(row=cont2, column=1).value = user.club_premier_id
        ws.cell(row=cont2, column=2).value = user.created
        total_score = Session.objects.filter(client_user_pk__club_premier_id__exact=user.club_premier_id).aggregate(Sum('high_score'))
        ws.cell(row=cont2, column=3).value = total_score['high_score__sum']
        completed_game =  Session.objects.filter(client_user_pk__club_premier_id__exact=user.club_premier_id, attempt=3).count()
        ws.cell(row=cont2, column=4).value = completed_game
        if user.accumulation is None:
            ws.cell(row=cont2, column=5).value = 0
            ws.cell(row=cont2, column=6).value = total_score['high_score__sum']
        else:
            ws.cell(row=cont2, column=5).value = user.accumulation
            accumulation=int(user.accumulation)*10
            if total_score['high_score__sum'] is None:
                ws.cell(row=cont2, column=6).value = 0 + accumulation
            else:
                ws.cell(row=cont2, column=6).value = int(total_score['high_score__sum'])+ accumulation

        cont2 = cont2 + 1
        percent = (cont/total)*100
        report.percent = percent
        report.save()
    nombre_archivo = "{name}.xlsx".format(name=report.name)
    wb.save('media/{name}.xlsx'.format(
            name=report.name))
    report.url = '/{name}.xlsx'.format(
            name=report.name)
    email_from = settings.EMAIL_HOST
    email_to = [email]
    msgworkshop = report.name
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    msg = EmailMessage(
        'Reporte Usuarios NBC',
        msgworkshop,
        email_from,
        [email_to],
    )
    msg.attach_file('media/{name}.xlsx'.format(
        name=report.name))
    msg.send()
    report.percent = 100
    report.save()
    print('correo enviado')

@app.task
def update_file(data):
    file = UpdateFile.objects.get(pk=data['id'])
    if file is None:
        return True
    read = file.file.open(mode="r")
    a = list(read)
    total_row = len(a)
    for index, r in enumerate(a):

        if index == 0:
            continue
        info = r.split(',')
        club_premier_id = info[0]
        accumulation = info[1]
        try:
            ClientUserControllers.create_user_file(club_premier_id, accumulation)
        except Exception as e:
            print(e)
    percent = (100 * index) / total_row
    file.percent = 100
    file.save()
